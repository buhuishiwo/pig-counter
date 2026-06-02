from __future__ import annotations

import base64
import os
import threading
import time
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

import cv2
import numpy as np
import pymysql
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile, Request
from ocr import recognize_farm_mark
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

SERVICE_DIR = Path(__file__).resolve().parent

load_dotenv(SERVICE_DIR / ".env")
STATIC_DIR = SERVICE_DIR / "static"
DEFAULT_MODEL_PATH = SERVICE_DIR / "model" / "pig_count.onnx"
DEFAULT_HOST = os.getenv("PIG_SERVICE_HOST", "0.0.0.0")
DEFAULT_PORT = int(os.getenv("PIG_SERVICE_PORT", "8866"))

# MySQL数据库配置
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "port": int(os.getenv("DB_PORT", "3306")),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD"),
    "database": os.getenv("DB_NAME", "pig_counter"),
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}

_OCR_API_KEY = os.getenv("OCR_API_KEY", "").strip()
_OCR_BASE_URL = os.getenv("OCR_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
_OCR_MODEL = os.getenv("OCR_MODEL", "qwen-vl-ocr")

_MODEL_LOCK = threading.Lock()
_MODEL_CACHE: dict[str, cv2.dnn.Net] = {}


@contextmanager
def get_db():
    """数据库连接上下文管理器"""
    conn = None
    try:
        conn = pymysql.connect(**DB_CONFIG)
        yield conn
    finally:
        if conn:
            conn.close()


class DetectionBox(BaseModel):
    x1: float
    y1: float
    x2: float
    y2: float
    confidence: float
    class_id: int
    class_name: str


class PredictResponse(BaseModel):
    success: bool
    model_path: str
    image_width: int
    image_height: int
    predicted_count: int
    detections: list[DetectionBox]
    processing_time_ms: float
    annotated_image: str
    record_id: int | None = None


class BatchPredictResponse(BaseModel):
    success: bool
    total_images: int
    total_pigs: int
    results: list[PredictResponse]


def resolve_model_path() -> Path:
    raw_path = os.getenv("PIG_MODEL_PATH")
    if raw_path:
        return Path(raw_path).expanduser().resolve()
    return DEFAULT_MODEL_PATH.resolve()


def get_model(model_path: Path) -> cv2.dnn.Net:
    if not model_path.exists():
        raise FileNotFoundError(f"模型文件不存在: {model_path}")

    model_key = str(model_path)
    with _MODEL_LOCK:
        if model_key not in _MODEL_CACHE:
            _MODEL_CACHE[model_key] = cv2.dnn.readNetFromONNX(str(model_path))
        return _MODEL_CACHE[model_key]


def decode_image(image_bytes: bytes) -> np.ndarray:
    image = cv2.imdecode(np.frombuffer(image_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
    if image is None:
        raise ValueError("无法解析上传图片")
    return image


def encode_image(image: np.ndarray) -> str:
    ok, buffer = cv2.imencode(".jpg", image)
    if not ok:
        raise ValueError("标注图编码失败")
    encoded = base64.b64encode(buffer.tobytes()).decode("utf-8")
    return f"data:image/jpeg;base64,{encoded}"


def create_thumbnail(image: np.ndarray, max_size: int = 320, quality: int = 60) -> str:
    height, width = image.shape[:2]
    scale = min(max_size / width, max_size / height)
    if scale < 1:
        new_width = int(width * scale)
        new_height = int(height * scale)
        thumbnail = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)
    else:
        thumbnail = image
    ok, buffer = cv2.imencode(".jpg", thumbnail, [cv2.IMWRITE_JPEG_QUALITY, quality])
    if not ok:
        raise ValueError("缩略图编码失败")
    encoded = base64.b64encode(buffer.tobytes()).decode("utf-8")
    return f"data:image/jpeg;base64,{encoded}"


def decode_base64_image(base64_str: str) -> np.ndarray:
    """从 base64 字符串解码图片"""
    if base64_str.startswith("data:image"):
        base64_str = base64_str.split(",")[1]
    img_data = base64.b64decode(base64_str)
    np_arr = np.frombuffer(img_data, np.uint8)
    return cv2.imdecode(np_arr, cv2.IMREAD_COLOR)


def draw_detections(image: np.ndarray, detections: list[DetectionBox]) -> np.ndarray:
    """在识别结果图上仅绘制数字序号，不绘制矩形框"""
    annotated = image.copy()
    for index, det in enumerate(detections, start=1):
        cx = int((det.x1 + det.x2) / 2)
        cy = int((det.y1 + det.y2) / 2)
        box_short = min(det.x2 - det.x1, det.y2 - det.y1)
        label = str(index)
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = max(2.0, min(box_short / 35.0, 4.0))
        thickness = max(3, int(font_scale * 2))
        (tw, th), _ = cv2.getTextSize(label, font, font_scale, thickness)
        tx = cx - tw // 2
        ty = cy + th // 2
        cv2.putText(annotated, label, (tx, ty), font, font_scale, (0, 0, 0), thickness, cv2.LINE_AA)
    return annotated


def letterbox(
    image: np.ndarray,
    new_size: int,
    color: tuple[int, int, int] = (114, 114, 114),
) -> tuple[np.ndarray, float, float, float]:
    height, width = image.shape[:2]
    scale = min(new_size / width, new_size / height)
    resized_width = int(round(width * scale))
    resized_height = int(round(height * scale))

    resized = cv2.resize(image, (resized_width, resized_height), interpolation=cv2.INTER_LINEAR)
    pad_w = new_size - resized_width
    pad_h = new_size - resized_height
    left = pad_w // 2
    right = pad_w - left
    top = pad_h // 2
    bottom = pad_h - top
    padded = cv2.copyMakeBorder(resized, top, bottom, left, right, cv2.BORDER_CONSTANT, value=color)
    return padded, scale, float(left), float(top)


def postprocess(
    output: np.ndarray,
    scale: float,
    pad_x: float,
    pad_y: float,
    original_width: int,
    original_height: int,
    conf_threshold: float,
    iou_threshold: float,
) -> list[DetectionBox]:
    predictions = output[0].T
    boxes_xywh: list[list[float]] = []
    confidences: list[float] = []
    class_ids: list[int] = []

    for row in predictions:
        if row.shape[0] <= 4:
            continue

        if row.shape[0] == 5:
            class_id = 0
            confidence = float(row[4])
        else:
            class_scores = row[4:]
            class_id = int(np.argmax(class_scores))
            confidence = float(class_scores[class_id])

        if confidence < conf_threshold:
            continue

        center_x, center_y, width, height = map(float, row[:4])
        x1 = (center_x - width / 2 - pad_x) / scale
        y1 = (center_y - height / 2 - pad_y) / scale
        x2 = (center_x + width / 2 - pad_x) / scale
        y2 = (center_y + height / 2 - pad_y) / scale

        x1 = max(0.0, min(x1, float(original_width)))
        y1 = max(0.0, min(y1, float(original_height)))
        x2 = max(0.0, min(x2, float(original_width)))
        y2 = max(0.0, min(y2, float(original_height)))
        width_box = max(0.0, x2 - x1)
        height_box = max(0.0, y2 - y1)
        if width_box <= 1 or height_box <= 1:
            continue

        boxes_xywh.append([x1, y1, width_box, height_box])
        confidences.append(confidence)
        class_ids.append(class_id)

    if not boxes_xywh:
        return []

    indices = cv2.dnn.NMSBoxes(boxes_xywh, confidences, conf_threshold, iou_threshold)
    if indices is None or len(indices) == 0:
        return []

    detections: list[DetectionBox] = []
    for raw_index in np.array(indices).reshape(-1):
        index = int(raw_index)
        x1, y1, width_box, height_box = boxes_xywh[index]
        detections.append(
            DetectionBox(
                x1=round(x1, 2),
                y1=round(y1, 2),
                x2=round(x1 + width_box, 2),
                y2=round(y1 + height_box, 2),
                confidence=round(confidences[index], 6),
                class_id=class_ids[index],
                class_name="pig",
            )
        )

    detections.sort(key=lambda item: item.confidence, reverse=True)
    return detections


def predict_image(
    image: np.ndarray,
    model_path: Path,
    conf_threshold: float,
    iou_threshold: float,
    imgsz: int,
) -> tuple[PredictResponse, np.ndarray]:
    started_at = time.perf_counter()
    image_height, image_width = image.shape[:2]
    processed, scale, pad_x, pad_y = letterbox(image, imgsz)
    blob = cv2.dnn.blobFromImage(processed, scalefactor=1 / 255.0, size=(imgsz, imgsz), swapRB=True, crop=False)

    model = get_model(model_path)
    with _MODEL_LOCK:
        model.setInput(blob)
        output = model.forward()

    detections = postprocess(
        output=output,
        scale=scale,
        pad_x=pad_x,
        pad_y=pad_y,
        original_width=image_width,
        original_height=image_height,
        conf_threshold=conf_threshold,
        iou_threshold=iou_threshold,
    )

    annotated = draw_detections(image, detections)
    response = PredictResponse(
        success=True,
        model_path=str(model_path),
        image_width=image_width,
        image_height=image_height,
        predicted_count=len(detections),
        detections=detections,
        processing_time_ms=round((time.perf_counter() - started_at) * 1000, 2),
        annotated_image=encode_image(annotated),
    )
    return response, annotated


app = FastAPI(
    title="Pig Count Service",
    description="猪只计数模型接口与测试页面",
    version="1.0.0",
)

# 添加异常处理器，处理请求体大小超过限制的错误
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from starlette.exceptions import HTTPException as StarletteHTTPException

@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request, exc):
    if exc.status_code == 413:
        return JSONResponse(
            status_code=413,
            content={"detail": "图片大小超过单次最大上传值！"}
        )
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail}
    )

# 处理请求验证错误
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(
        status_code=422,
        content={"detail": exc.errors()}  # 👈 保留原始错误
    )




@app.on_event("startup")
async def startup_event() -> None:
    model_path = resolve_model_path()
    if not model_path.exists():
        raise RuntimeError(f"模型文件不存在: {model_path}")
    get_model(model_path)




@app.get("/api/health")
async def health() -> dict[str, Any]:
    model_path = resolve_model_path()
    return {
        "status": "healthy" if model_path.exists() else "unhealthy",
        "model_path": str(model_path),
        "model_exists": model_path.exists(),
        "runtime": "opencv-dnn",
    }


@app.get("/api/config")
async def config() -> dict[str, Any]:
    model_path = resolve_model_path()
    return {
        "service_name": "pig-count-service",
        "model_path": str(model_path),
        "default_conf_threshold": 0.25,
        "default_iou_threshold": 0.45,
        "default_imgsz": 960,
    }


@app.post("/api/predict-batch", response_model=BatchPredictResponse)
async def predict_batch(
    files: list[UploadFile] = File(...),
    farm_id: int | None = Form(default=None),
    conf_threshold: float = Form(default=0.25),
    iou_threshold: float = Form(default=0.45),
    imgsz: int = Form(default=960),
) -> BatchPredictResponse:
    if not files:
        raise HTTPException(status_code=400, detail="至少需要上传一张图片")

    for file in files:
        if file.content_type and not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail=f"文件 {file.filename} 不是图片文件")

    if not 0 <= conf_threshold <= 1:
        raise HTTPException(status_code=400, detail="conf_threshold 必须在 0 到 1 之间")
    if not 0 <= iou_threshold <= 1:
        raise HTTPException(status_code=400, detail="iou_threshold 必须在 0 到 1 之间")
    if imgsz <= 0:
        raise HTTPException(status_code=400, detail="imgsz 必须大于 0")

    model_path = resolve_model_path()
    results = []
    total_pigs = 0

    try:
        for file in files:
            image_bytes = await file.read()
            image = decode_image(image_bytes)
            result, annotated_image = predict_image(
                image=image,
                model_path=model_path,
                conf_threshold=conf_threshold,
                iou_threshold=iou_threshold,
                imgsz=imgsz,
            )
            
            # 计算平均置信度
            avg_confidence = (
                sum(d.confidence for d in result.detections) / len(result.detections)
                if result.detections else 0
            )
            boxes_data = [
                {"x1": d.x1, "y1": d.y1, "x2": d.x2, "y2": d.y2,
                 "score": d.confidence, "class_id": d.class_id, "class_name": d.class_name}
                for d in result.detections
            ] if result.detections else []

            # 保存识别记录到数据库
            record_id = await save_detection_record(
                farm_id=farm_id,
                image_name=file.filename or "unknown.jpg",
                predicted_count=result.predicted_count,
                processing_time_ms=result.processing_time_ms,
                annotated_image=annotated_image,
                confidence=avg_confidence,
                boxes=boxes_data,
                original_image=image,
            )

            # 设置返回结果的 record_id
            result = result.model_copy(update={"record_id": record_id})
            results.append(result)
            total_pigs += result.predicted_count
        
        return BatchPredictResponse(
            success=True,
            total_images=len(files),
            total_pigs=total_pigs,
            results=results
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"推理失败: {exc}") from exc


async def save_detection_record(
    farm_id: int | None,
    image_name: str,
    predicted_count: int,
    processing_time_ms: float,
    annotated_image: np.ndarray,
    confidence: float = 0,
    boxes: list[dict] | None = None,
    original_image: np.ndarray | None = None,
) -> int | None:
    """保存识别记录到数据库（存储压缩缩略图），返回 record_id"""
    try:
        import json as _json
        thumbnail_base64 = create_thumbnail(annotated_image, max_size=800, quality=90)
        orig_thumbnail = create_thumbnail(original_image, max_size=800, quality=85) if original_image is not None else None
        boxes_json = _json.dumps(boxes, ensure_ascii=False) if boxes else None
        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute(
                    """INSERT INTO detection_records
                        (farm_id, image_name, predicted_count, processing_time_ms, confidence, boxes, annotated_image, original_image, created_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())""",
                    (farm_id, image_name, predicted_count, processing_time_ms, confidence, boxes_json, thumbnail_base64, orig_thumbnail)
                )
                conn.commit()
                return cursor.lastrowid
    except Exception as exc:
        print(f"保存识别记录失败: {exc}")
        return None


# ============================================================
# OCR 标记识别 API
# ============================================================

class OCRMarkResponse(BaseModel):
    success: bool
    text: str | None = None
    source: str | None = None
    confidence: float | None = None
    source_quality: str | None = None
    suggestions: list[dict[str, Any]] = []


@app.post("/api/ocr/farm-mark", response_model=OCRMarkResponse)
async def ocr_farm_mark(file: UploadFile = File(...)) -> OCRMarkResponse:
    if file.content_type and not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="只支持图片文件")

    image_bytes = await file.read()
    if len(image_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="图片大小超过 10MB 限制")

    try:
        image = decode_image(image_bytes)
        result = recognize_farm_mark(image)
    except ValueError:
        raise HTTPException(status_code=400, detail="无法解析图片，请确认文件格式正确")
    except Exception:
        return OCRMarkResponse(success=False, text=None, source=None, suggestions=[])

    text = result.get("text")
    source = result.get("source")
    confidence = result.get("confidence")
    source_quality = result.get("source_quality")

    suggestions: list[dict[str, Any]] = []
    if text:
        try:
            with get_db() as conn:
                with conn.cursor() as cursor:
                    parts = [p for p in text.replace("-", " ").replace("/", " ").split() if len(p) >= 2]
                    if parts:
                        conditions = []
                        params = []
                        for part in parts:
                            conditions.append("name LIKE %s")
                            params.append(f"%{part}%")
                        conditions.append("%s LIKE CONCAT('%%', name, '%%')")
                        params.append(text)

                        sql = (
                            "SELECT DISTINCT id, name FROM pig_farms"
                            f" WHERE {' OR '.join(conditions)} LIMIT 5"
                        )
                        cursor.execute(sql, params)
                        for row in cursor.fetchall():
                            suggestions.append({"farm_id": row["id"], "name": row["name"]})
        except Exception:
            pass

    return OCRMarkResponse(
        success=text is not None,
        text=text,
        source=source,
        confidence=confidence,
        source_quality=source_quality,
        suggestions=suggestions,
    )


# ============================================================
# 猪场管理API
# ============================================================

class PigFarmCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100, description="猪场名称")


class PigFarmUpdate(BaseModel):
    name: str = Field(..., min_length=1, max_length=100, description="猪场名称")


class PigFarmResponse(BaseModel):
    id: int
    name: str
    created_at: datetime

    class Config:
        from_attributes = True


class PigFarmListResponse(BaseModel):
    success: bool
    data: list[PigFarmResponse]


class PigFarmDetailResponse(BaseModel):
    success: bool
    data: PigFarmResponse | None


class PigFarmMessageResponse(BaseModel):
    success: bool
    message: str


@app.get("/api/farms", response_model=PigFarmListResponse)
async def get_farms() -> PigFarmListResponse:
    """获取所有猪场列表"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT id, name, created_at FROM pig_farms ORDER BY created_at DESC"
                )
                farms = cursor.fetchall()
                return PigFarmListResponse(
                    success=True,
                    data=[
                        PigFarmResponse(
                            id=farm["id"],
                            name=farm["name"],
                            created_at=farm["created_at"],
                        )
                        for farm in farms
                    ],
                )
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取猪场列表失败: {exc}"
        ) from exc


@app.get("/api/farms/{farm_id}", response_model=PigFarmDetailResponse)
async def get_farm(farm_id: int) -> PigFarmDetailResponse:
    """获取单个猪场信息"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute(
                    "SELECT id, name, created_at FROM pig_farms WHERE id = %s",
                    (farm_id,),
                )
                farm = cursor.fetchone()
                if not farm:
                    raise HTTPException(status_code=404, detail="猪场不存在")
                return PigFarmDetailResponse(
                    success=True,
                    data=PigFarmResponse(
                        id=farm["id"],
                        name=farm["name"],
                        created_at=farm["created_at"],
                    ),
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取猪场信息失败: {exc}"
        ) from exc


@app.post("/api/farms", response_model=PigFarmDetailResponse)
async def create_farm(farm: PigFarmCreate) -> PigFarmDetailResponse:
    """创建新猪场"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 检查名称是否已存在
                cursor.execute(
                    "SELECT id FROM pig_farms WHERE name = %s", (farm.name,)
                )
                if cursor.fetchone():
                    raise HTTPException(status_code=400, detail="猪场名称已存在")

                cursor.execute(
                    "INSERT INTO pig_farms (name, created_at) VALUES (%s, NOW())",
                    (farm.name,),
                )
                conn.commit()
                farm_id = cursor.lastrowid

                cursor.execute(
                    "SELECT id, name, created_at FROM pig_farms WHERE id = %s",
                    (farm_id,),
                )
                new_farm = cursor.fetchone()
                return PigFarmDetailResponse(
                    success=True,
                    data=PigFarmResponse(
                        id=new_farm["id"],
                        name=new_farm["name"],
                        created_at=new_farm["created_at"],
                    ),
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"创建猪场失败: {exc}"
        ) from exc


@app.put("/api/farms/{farm_id}", response_model=PigFarmDetailResponse)
async def update_farm(farm_id: int, farm: PigFarmUpdate) -> PigFarmDetailResponse:
    """更新猪场信息"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 检查猪场是否存在
                cursor.execute(
                    "SELECT id FROM pig_farms WHERE id = %s", (farm_id,)
                )
                if not cursor.fetchone():
                    raise HTTPException(status_code=404, detail="猪场不存在")

                # 检查新名称是否与其他猪场冲突
                cursor.execute(
                    "SELECT id FROM pig_farms WHERE name = %s AND id != %s",
                    (farm.name, farm_id),
                )
                if cursor.fetchone():
                    raise HTTPException(status_code=400, detail="猪场名称已存在")

                cursor.execute(
                    "UPDATE pig_farms SET name = %s WHERE id = %s",
                    (farm.name, farm_id),
                )
                conn.commit()

                cursor.execute(
                    "SELECT id, name, created_at FROM pig_farms WHERE id = %s",
                    (farm_id,),
                )
                updated_farm = cursor.fetchone()
                return PigFarmDetailResponse(
                    success=True,
                    data=PigFarmResponse(
                        id=updated_farm["id"],
                        name=updated_farm["name"],
                        created_at=updated_farm["created_at"],
                    ),
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"更新猪场失败: {exc}"
        ) from exc


@app.delete("/api/farms/{farm_id}", response_model=PigFarmMessageResponse)
async def delete_farm(farm_id: int) -> PigFarmMessageResponse:
    """删除猪场"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 检查猪场是否存在
                cursor.execute(
                    "SELECT id FROM pig_farms WHERE id = %s", (farm_id,)
                )
                if not cursor.fetchone():
                    raise HTTPException(status_code=404, detail="猪场不存在")

                cursor.execute("DELETE FROM pig_farms WHERE id = %s", (farm_id,))
                conn.commit()
                return PigFarmMessageResponse(
                    success=True, message="猪场删除成功"
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"删除猪场失败: {exc}"
        ) from exc


# ============================================================
# 识别记录统计API
# ============================================================

class DetectionRecordResponse(BaseModel):
    id: int
    farm_id: int | None
    farm_name: str | None
    image_name: str
    predicted_count: int
    processing_time_ms: float
    boxes: list | None = None
    annotated_image: str | None = None
    created_at: datetime


class UpdateDetectionRecordRequest(BaseModel):
    annotated_image: str | None = None
    original_image: str | None = None
    predicted_count: int | None = None
    boxes: list | None = None


class DetectionRecordListResponse(BaseModel):
    success: bool
    data: list[DetectionRecordResponse]
    total: int


class DetectionStatsResponse(BaseModel):
    success: bool
    data: dict[str, Any]


@app.get("/api/detection-records", response_model=DetectionRecordListResponse)
async def get_detection_records(
    farm_id: int | None = None,
    page: int = 1,
    page_size: int = 20
) -> DetectionRecordListResponse:
    """获取识别记录列表"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 构建查询条件
                where_clause = ""
                params = []
                if farm_id:
                    where_clause = "WHERE dr.farm_id = %s"
                    params.append(farm_id)
                
                # 查询总数
                count_sql = f"SELECT COUNT(*) as total FROM detection_records dr {where_clause}"
                cursor.execute(count_sql, params)
                total = cursor.fetchone()["total"]
                
                # 查询记录
                offset = (page - 1) * page_size
                sql = f"""
                    SELECT dr.id, dr.farm_id, pf.name as farm_name, dr.image_name,
                           dr.predicted_count, dr.processing_time_ms, dr.created_at
                    FROM detection_records dr
                    LEFT JOIN pig_farms pf ON dr.farm_id = pf.id
                    {where_clause}
                    ORDER BY dr.created_at DESC
                    LIMIT %s OFFSET %s
                """
                cursor.execute(sql, params + [page_size, offset])
                records = cursor.fetchall()
                
                return DetectionRecordListResponse(
                    success=True,
                    data=[
                        DetectionRecordResponse(
                            id=r["id"],
                            farm_id=r["farm_id"],
                            farm_name=r["farm_name"],
                            image_name=r["image_name"],
                            predicted_count=r["predicted_count"],
                            processing_time_ms=r["processing_time_ms"],
                            created_at=r["created_at"]
                        )
                        for r in records
                    ],
                    total=total
                )
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取识别记录失败: {exc}"
        ) from exc


class DetectionRecordWithImageResponse(BaseModel):
    id: int
    farm_id: int | None
    farm_name: str | None
    image_name: str
    predicted_count: int
    processing_time_ms: float
    confidence: float = 0
    annotated_image: str | None  # base64 缩略图
    created_at: datetime
 
 
class DetectionRecordWithImageListResponse(BaseModel):
    success: bool
    data: list[DetectionRecordWithImageResponse]
    total: int
    page: int
    page_size: int


@app.get("/api/detection-records/with-images", response_model=DetectionRecordWithImageListResponse)
async def get_detection_records_with_images(
    farm_id: int | None = None,
    page: int = 1,
    page_size: int = 12,
    start_date: str | None = None,
    end_date: str | None = None,
    keyword: str | None = None,
) -> DetectionRecordWithImageListResponse:

    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    # 记录请求参数
    logger.info(f"Received request to /api/detection-records/with-images with params: farm_id={farm_id}, page={page}, page_size={page_size}")
    
    # 验证参数
    if page < 1:
        logger.error(f"Invalid page parameter: {page}")
        raise HTTPException(status_code=422, detail="page must be greater than 0")
    if page_size < 1 or page_size > 50:
        logger.error(f"Invalid page_size parameter: {page_size}")
        raise HTTPException(status_code=422, detail="page_size must be between 1 and 50")
    if farm_id is not None and not isinstance(farm_id, int):
        logger.error(f"Invalid farm_id parameter type: {type(farm_id)}")
        raise HTTPException(status_code=422, detail="farm_id must be an integer")
    
    logger.info(f"Parameters validated successfully")
    page_size = min(page_size, 50)
    try:
        logger.info("Starting database operations")
        with get_db() as conn:
            logger.info("Database connection established")
            with conn.cursor() as cursor:
                conditions: list[str] = []
                params: list = []
                if farm_id is not None:
                    conditions.append("dr.farm_id = %s")
                    params.append(farm_id)
                if start_date:
                    conditions.append("dr.created_at >= %s")
                    params.append(start_date)
                if end_date:
                    conditions.append("dr.created_at <= %s")
                    params.append(end_date + " 23:59:59")
                if keyword:
                    conditions.append("dr.image_name LIKE %s")
                    params.append(f"%{keyword}%")
                where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""

                # 总数
                count_sql = f"SELECT COUNT(*) AS total FROM detection_records dr {where_clause}"
                logger.info(f"Executing count query: {count_sql} with params: {params}")
                cursor.execute(count_sql, params)
                total = cursor.fetchone()["total"]
                logger.info(f"Total records found: {total}")

                offset = (page - 1) * page_size
                logger.info(f"Calculated offset: {offset}")
                
                select_sql = f"""
                    SELECT
                        dr.id, dr.farm_id, pf.name AS farm_name,
                        dr.image_name, dr.predicted_count,
                        dr.processing_time_ms, dr.confidence, dr.annotated_image, dr.created_at
                    FROM detection_records dr
                    LEFT JOIN pig_farms pf ON dr.farm_id = pf.id
                    {where_clause}
                    ORDER BY dr.created_at DESC
                    LIMIT %s OFFSET %s
                """
                logger.info(f"Executing select query with params: {params + [page_size, offset]}")
                cursor.execute(select_sql, params + [page_size, offset])
                records = cursor.fetchall()
                logger.info(f"Fetched {len(records)} records")

                response_data = [
                    DetectionRecordWithImageResponse(
                        id=r["id"],
                        farm_id=r["farm_id"],
                        farm_name=r["farm_name"],
                        image_name=r["image_name"],
                        predicted_count=r["predicted_count"],
                        processing_time_ms=r["processing_time_ms"],
                        confidence=float(r.get("confidence") or 0),
                        annotated_image=r["annotated_image"],
                        created_at=r["created_at"],
                    )
                    for r in records
                ]
                
                logger.info(f"Prepared response with {len(response_data)} items")
                
                return DetectionRecordWithImageListResponse(
                    success=True,
                    data=response_data,
                    total=total,
                    page=page,
                    page_size=page_size,
                )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error in get_detection_records_with_images: {exc}", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"获取识别记录失败: {exc}"
        ) from exc


@app.get("/api/detection-records/{record_id}", response_model=PredictResponse)
async def get_detection_record_detail(record_id: int) -> PredictResponse:
    """获取识别记录详细信息"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 查询记录
                cursor.execute(
                    """
                    SELECT dr.id, dr.farm_id, dr.image_name, dr.predicted_count, 
                           dr.processing_time_ms, dr.annotated_image
                    FROM detection_records dr
                    WHERE dr.id = %s
                    """,
                    (record_id,)
                )
                record = cursor.fetchone()
                
                if not record:
                    raise HTTPException(status_code=404, detail="识别记录不存在")
                
                # 解析标注图片和检测结果
                # 注意：这里我们没有存储原始的检测框数据，所以返回空数组
                # 在实际应用中，你可能需要修改数据库结构，存储检测框数据
                detections = []
                
                return PredictResponse(
                    success=True,
                    model_path=str(resolve_model_path()),
                    image_width=0,  # 这里可以从存储的信息中获取，或者设为0
                    image_height=0, # 这里可以从存储的信息中获取，或者设为0
                    predicted_count=record["predicted_count"],
                    detections=detections,
                    processing_time_ms=record["processing_time_ms"],
                    annotated_image=record["annotated_image"]
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取识别记录详情失败: {exc}"
        ) from exc


@app.put("/api/detection-records/{record_id}", response_model=DetectionRecordResponse)
async def update_detection_record(
    record_id: int,
    update_data: UpdateDetectionRecordRequest
) -> DetectionRecordResponse:
    """更新识别记录（仅允许更新 annotated_image、predicted_count 和 boxes）"""
    try:
        import json as _json
        with get_db() as conn:
            with conn.cursor() as cursor:
                updates = []
                params = []

                if update_data.boxes is not None:
                    # 优先用前端发来的原图，其次用数据库原图，最后用标注图
                    img_array = None
                    if update_data.original_image:
                        img_array = decode_base64_image(update_data.original_image)
                    if img_array is None:
                        cursor.execute("SELECT original_image FROM detection_records WHERE id = %s", (record_id,))
                        row = cursor.fetchone()
                        if row and row["original_image"]:
                            orig_b64 = row["original_image"]
                            if orig_b64.startswith("data:"):
                                orig_b64 = orig_b64.split(",")[1]
                            import base64 as _base64
                            orig_bytes = _base64.b64decode(orig_b64)
                            img_array = cv2.imdecode(np.frombuffer(orig_bytes, np.uint8), cv2.IMREAD_COLOR)
                    if img_array is None and update_data.annotated_image is not None:
                        img_array = decode_base64_image(update_data.annotated_image)

                    if img_array is not None:
                        boxes_list = update_data.boxes
                        for idx, box in enumerate(boxes_list, start=1):
                            x1 = int(box.get("x1", 0))
                            y1 = int(box.get("y1", 0))
                            x2 = int(box.get("x2", 0))
                            y2 = int(box.get("y2", 0))
                            cx = (x1 + x2) // 2
                            cy = (y1 + y2) // 2
                            box_short = min(x2 - x1, y2 - y1)
                            font = cv2.FONT_HERSHEY_SIMPLEX
                            font_scale = max(2.0, min(box_short / 35.0, 4.0))
                            thickness = max(3, int(font_scale * 2))
                            label = str(idx)
                            (tw, th), _ = cv2.getTextSize(label, font, font_scale, thickness)
                            tx = cx - tw // 2
                            ty = cy + th // 2
                            cv2.putText(img_array, label, (tx, ty), font, font_scale, (0, 0, 0), thickness, cv2.LINE_AA)
                        thumbnail_base64 = create_thumbnail(img_array, max_size=800, quality=90)
                        updates.append("annotated_image = %s")
                        params.append(thumbnail_base64)

                    updates.append("boxes = %s")
                    params.append(_json.dumps(update_data.boxes))

                if update_data.annotated_image is not None and update_data.boxes is None:
                    updates.append("annotated_image = %s")
                    params.append(update_data.annotated_image)
                if update_data.predicted_count is not None:
                    updates.append("predicted_count = %s")
                    params.append(update_data.predicted_count)

                if not updates:
                    raise HTTPException(status_code=400, detail="没有提供要更新的字段")

                params.append(record_id)
                cursor.execute(
                    f"UPDATE detection_records SET {', '.join(updates)} WHERE id = %s",
                    params
                )
                conn.commit()

                if cursor.rowcount == 0:
                    raise HTTPException(status_code=404, detail="识别记录不存在")

                cursor.execute(
                    """
                    SELECT dr.id, dr.farm_id, pf.name AS farm_name, dr.image_name,
                           dr.predicted_count, dr.processing_time_ms, dr.annotated_image,
                           dr.boxes, dr.created_at
                    FROM detection_records dr
                    LEFT JOIN pig_farms pf ON dr.farm_id = pf.id
                    WHERE dr.id = %s
                    """,
                    (record_id,)
                )
                record = cursor.fetchone()
                boxes = None
                if record["boxes"]:
                    try:
                        boxes = _json.loads(record["boxes"])
                    except Exception:
                        boxes = None

                return DetectionRecordResponse(
                    id=record["id"],
                    farm_id=record["farm_id"],
                    farm_name=record["farm_name"],
                    image_name=record["image_name"],
                    predicted_count=record["predicted_count"],
                    processing_time_ms=record["processing_time_ms"],
                    boxes=boxes,
                    annotated_image=record["annotated_image"],
                    created_at=record["created_at"],
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"更新识别记录失败: {exc}"
        ) from exc


@app.get("/api/detection-stats", response_model=DetectionStatsResponse)
async def get_detection_stats(
    farm_id: int | None = None
) -> DetectionStatsResponse:
    """获取识别统计信息"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 构建查询条件
                where_clause = ""
                params = []
                if farm_id:
                    where_clause = "WHERE farm_id = %s"
                    params.append(farm_id)
                
                # 总识别图片数
                cursor.execute(
                    f"SELECT COUNT(*) as total_images FROM detection_records {where_clause}",
                    params
                )
                total_images = cursor.fetchone()["total_images"]
                
                # 总识别猪数量
                cursor.execute(
                    f"SELECT COALESCE(SUM(predicted_count), 0) as total_pigs FROM detection_records {where_clause}",
                    params
                )
                total_pigs = cursor.fetchone()["total_pigs"]
                
                # 今日识别数量
                today_where = f"{where_clause} AND DATE(created_at) = CURDATE()" if where_clause else "WHERE DATE(created_at) = CURDATE()"
                cursor.execute(
                    f"SELECT COUNT(*) as today_images, COALESCE(SUM(predicted_count), 0) as today_pigs FROM detection_records {today_where}",
                    params
                )
                today_stats = cursor.fetchone()
                
                # 平均处理时间
                cursor.execute(
                    f"SELECT AVG(processing_time_ms) as avg_time FROM detection_records {where_clause}",
                    params
                )
                avg_time = cursor.fetchone()["avg_time"] or 0
                
                return DetectionStatsResponse(
                    success=True,
                    data={
                        "total_images": total_images,
                        "total_pigs": int(total_pigs),
                        "today_images": today_stats["today_images"],
                        "today_pigs": int(today_stats["today_pigs"]),
                        "avg_processing_time_ms": round(avg_time, 2)
                    }
                )
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取统计数据失败: {exc}"
        ) from exc

class FarmStatsItem(BaseModel):
    farm_id: int | None
    farm_name: str | None
    total_images: int
    total_pigs: int
    today_images: int
    today_pigs: int
    avg_processing_time_ms: float
    last_detection_at: datetime | None
 
 
class FarmStatsListResponse(BaseModel):
    success: bool
    data: list[FarmStatsItem]
 
 
@app.get("/api/detection-stats/by-farm", response_model=FarmStatsListResponse)
async def get_stats_by_farm() -> FarmStatsListResponse:
    """获取按猪场分组的识别统计数据（包含无猪场归属的记录）"""
    try:
        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT
                        dr.farm_id,
                        pf.name AS farm_name,
                        COUNT(*) AS total_images,
                        COALESCE(SUM(dr.predicted_count), 0) AS total_pigs,
                        SUM(DATE(dr.created_at) = CURDATE()) AS today_images,
                        COALESCE(SUM(CASE WHEN DATE(dr.created_at) = CURDATE() THEN dr.predicted_count ELSE 0 END), 0) AS today_pigs,
                        COALESCE(AVG(dr.processing_time_ms), 0) AS avg_processing_time_ms,
                        MAX(dr.created_at) AS last_detection_at
                    FROM detection_records dr
                    LEFT JOIN pig_farms pf ON dr.farm_id = pf.id
                    GROUP BY dr.farm_id, pf.name
                    ORDER BY total_images DESC
                    """
                )
                rows = cursor.fetchall()
 
                return FarmStatsListResponse(
                    success=True,
                    data=[
                        FarmStatsItem(
                            farm_id=r["farm_id"],
                            farm_name=r["farm_name"],
                            total_images=r["total_images"],
                            total_pigs=int(r["total_pigs"]),
                            today_images=int(r["today_images"] or 0),
                            today_pigs=int(r["today_pigs"]),
                            avg_processing_time_ms=round(float(r["avg_processing_time_ms"]), 2),
                            last_detection_at=r["last_detection_at"],
                        )
                        for r in rows
                    ],
                )
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取猪场统计数据失败: {exc}"
        ) from exc


class TimeSeriesDataItem(BaseModel):
    date: str
    images: int
    pigs: int


class TimeSeriesResponse(BaseModel):
    success: bool
    data: list[TimeSeriesDataItem]
    granularity: str


@app.get("/api/detection-stats/time-series", response_model=TimeSeriesResponse)
async def get_time_series_stats(
    granularity: str = "day",  # day, month 或 year
    farm_id: int | None = None,
    days: int = 30
) -> TimeSeriesResponse:
    """获取按时间粒度统计的数据"""
    try:
        if granularity not in ["day", "month", "year"]:
            raise HTTPException(status_code=422, detail="granularity must be 'day', 'month' or 'year'")
        
        with get_db() as conn:
            with conn.cursor() as cursor:
                # 构建查询条件
                where_clause = ""
                params = []
                if farm_id is not None:
                    where_clause = "WHERE farm_id = %s"
                    params.append(farm_id)
                
                # 根据粒度构建日期格式
                if granularity == "day":
                    date_format = "DATE(created_at)"
                    time_condition = f"created_at >= DATE_SUB(NOW(), INTERVAL {days} DAY)"
                elif granularity == "month":
                    date_format = "DATE_FORMAT(created_at, '%%Y-%%m')"
                    time_condition = f"created_at >= DATE_SUB(NOW(), INTERVAL {days} DAY)"
                else:  # year
                    date_format = "DATE_FORMAT(created_at, '%%Y')"
                    time_condition = "1=1"
                
                # 构建完整的WHERE子句
                if where_clause:
                    where_clause = f"{where_clause} AND {time_condition}"
                else:
                    where_clause = f"WHERE {time_condition}"
                
                # 执行查询
                cursor.execute(
                    f"""
                    SELECT
                        {date_format} AS date,
                        COUNT(*) AS images,
                        COALESCE(SUM(predicted_count), 0) AS pigs
                    FROM detection_records
                    {where_clause}
                    GROUP BY date
                    ORDER BY date
                    """,
                    params
                )
                rows = cursor.fetchall()
                
                # 构建响应数据
                data = [
                    TimeSeriesDataItem(
                        date=str(r["date"]),
                        images=r["images"],
                        pigs=int(r["pigs"])
                    )
                    for r in rows
                ]
                
                return TimeSeriesResponse(
                    success=True,
                    data=data,
                    granularity=granularity
                )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500, detail=f"获取时间序列数据失败: {exc}"
        ) from exc

# ============================================================
# 文件夹批量上传 API（目录结构即身份）
# ============================================================

from io import BytesIO as _BytesIO
from collections import OrderedDict as _OrderedDict


class BatchUploadResponse(BaseModel):
    success: bool
    batch_name: str
    units: list[dict]
    total_photos: int
    total_pigs: int
    excel_base64: str = ""


@app.post("/api/batch/upload")
async def batch_upload(files: list[UploadFile] = File(...), farm_id: int | None = Form(default=None), file_paths: list[str] = Form(default=[])) -> dict:
    """
    接收带路径的多文件（前端 webkitdirectory 上传）。
    从文件相对路径解析 批次/单元/栏舍 三层结构。
    逐张推理，按单元汇总，返回结果 + Excel。
    """
    if not files:
        raise HTTPException(status_code=400, detail="至少需要上传一张图片")

    # 1. 解析路径，建树
    MAX_BATCH_FILES = 500
    if len(files) > MAX_BATCH_FILES:
        raise HTTPException(status_code=400, detail=f"单次最多上传 {MAX_BATCH_FILES} 张图片")

    batch_name = ""
    unit_order: list[str] = []
    units: dict[str, list[dict]] = _OrderedDict()  # unit_name → [{name, file}]

    for idx, f in enumerate(files):
        # 优先使用前端传来的完整路径
        if idx < len(file_paths) and file_paths[idx]:
            path = file_paths[idx].replace("\\", "/")
        else:
            path = (f.filename or "").replace("\\", "/")
        parts = [p for p in path.split("/") if p]

        if len(parts) < 2:
            continue  # 跳过根目录下的散文件

        if not batch_name:
            batch_name = parts[0]

        unit_name = parts[1] if len(parts) >= 2 else ""
        file_name = parts[-1]

        # 跳过栏舍号标识照
        name_no_ext = file_name.rsplit(".", 1)[0] if "." in file_name else file_name
        if name_no_ext == "栏舍号":
            continue

        if not f.content_type or not f.content_type.startswith("image/"):
            continue

        if unit_name not in units:
            units[unit_name] = []
            unit_order.append(unit_name)

        units[unit_name].append({"name": file_name, "file": f, "full_path": path})

    if not units:
        raise HTTPException(status_code=400, detail="未找到有效的图片文件（请确认目录结构：批次/单元/栏舍.jpg）")

    # 2. 逐张推理
    model_path = resolve_model_path()
    total_photos = 0
    total_pigs = 0
    results: list[dict] = []
    image_paths: list[str] = []  # 用于 Excel 插图

    # 创建临时目录存放原图
    import tempfile
    tmp_dir = Path(tempfile.mkdtemp(prefix="pigcount_batch_"))

    for unit_name in unit_order:
        pens = units[unit_name]
        unit_pigs = 0
        unit_results: list[dict] = []

        for pen in pens:
            result = None
            record_id = None
            try:
                image_bytes = await pen["file"].read()

                # 保存原图到临时目录
                img_path = tmp_dir / f"{unit_name}_{pen['name']}"
                img_path.write_bytes(image_bytes)

                image = decode_image(image_bytes)
                result, _annotated = predict_image(
                    image=image,
                    model_path=model_path,
                    conf_threshold=0.25,
                    iou_threshold=0.45,
                    imgsz=960,
                )
                count = result.predicted_count

                # 保存标注图到临时目录，用于 Excel 插图（带数字标注）
                import cv2 as _cv2
                ann_path = tmp_dir / f"ann_{unit_name}_{pen['name']}"
                _cv2.imwrite(str(ann_path), _annotated)
                image_paths.append(str(ann_path))
                avg_conf = (
                    sum(d.confidence for d in result.detections) / len(result.detections)
                    if result.detections else 0
                )
                boxes_for_db = [
                    {"x1": d.x1, "y1": d.y1, "x2": d.x2, "y2": d.y2,
                     "score": d.confidence, "class_id": d.class_id, "class_name": d.class_name}
                    for d in result.detections
                ] if result.detections else []
                record_id = await save_detection_record(
                    farm_id=farm_id,
                    image_name=pen["full_path"],
                    predicted_count=count,
                    processing_time_ms=result.processing_time_ms,
                    annotated_image=_annotated,
                    confidence=avg_conf,
                    boxes=boxes_for_db,
                    original_image=image,
                )
            except Exception:
                count = 0

            pen_data = {
                "pen_name": pen["name"],
                "pig_count": count,
                "processing_time_ms": result.processing_time_ms if result is not None else 0,
                "annotated_image": create_thumbnail(_annotated, max_size=480, quality=70) if result is not None and _annotated is not None else None,
                "record_id": record_id,
                "image_width": result.image_width if result is not None else 0,
                "image_height": result.image_height if result is not None else 0,
                "confidence": (
                    sum(d.confidence for d in result.detections) / len(result.detections)
                    if result and result.detections else 0
                ),
                "boxes": [
                    {"x1": d.x1, "y1": d.y1, "x2": d.x2, "y2": d.y2,
                     "score": d.confidence, "class_id": d.class_id, "class_name": d.class_name}
                    for d in (result.detections if result else [])
                ],
            }
            unit_results.append(pen_data)
            unit_pigs += count
            total_photos += 1
            total_pigs += count

        results.append({
            "unit_name": unit_name,
            "pens": unit_results,
            "subtotal": unit_pigs,
        })

    # 3. 生成 Excel（承储单位固定为公司名称）
    farm_name = "乐清市华统牧业有限公司"
    excel_b64 = _build_batch_excel(batch_name, results, farm_name, image_paths)

    # 5. 清理临时文件
    import shutil
    try:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception:
        pass

    return {
        "success": True,
        "batch_name": batch_name,
        "units": results,
        "total_photos": total_photos,
        "total_pigs": total_pigs,
        "excel_base64": excel_b64,
    }


@app.post("/api/batch/regenerate-excel")
async def regenerate_batch_excel(request: dict):
    """根据最新的 annotated_image 重新生成 Excel"""
    try:
        batch_name = request.get("batch_name", "批次统计")
        units = request.get("units", [])
        if not units:
            raise HTTPException(status_code=400, detail="缺少 units 数据")

        # 收集所有 record_id，批量查询最新的 annotated_image
        record_ids = []
        for unit in units:
            for pen in unit.get("pens", []):
                if pen.get("record_id"):
                    record_ids.append(pen["record_id"])

        if not record_ids:
            raise HTTPException(status_code=400, detail="无有效记录")

        placeholders = ",".join(["%s"] * len(record_ids))
        with get_db() as conn:
            with conn.cursor() as cursor:
                cursor.execute(
                    f"SELECT id, annotated_image FROM detection_records WHERE id IN ({placeholders})",
                    record_ids
                )
                rows = cursor.fetchall()
                image_map = {r["id"]: r["annotated_image"] for r in rows}

        # 用最新的 annotated_image 生成临时图片文件
        import tempfile, cv2 as _cv2
        tmp_dir = Path(tempfile.mkdtemp(prefix="pigcount_excel_"))
        image_paths = []
        flat_pens = []
        for unit in units:
            for pen in unit.get("pens", []):
                flat_pens.append(pen)
                rid = pen.get("record_id")
                ann_b64 = image_map.get(rid)
                if ann_b64:
                    if ann_b64.startswith("data:"):
                        ann_b64 = ann_b64.split(",")[1]
                    import base64 as _base64
                    img_bytes = _base64.b64decode(ann_b64)
                    img_array = _cv2.imdecode(np.frombuffer(img_bytes, np.uint8), _cv2.IMREAD_COLOR)
                    if img_array is not None:
                        ann_path = tmp_dir / f"ann_{rid}.jpg"
                        _cv2.imwrite(str(ann_path), img_array)
                        image_paths.append(str(ann_path))
                        continue
                image_paths.append(None)

        farm_name = "乐清市华统牧业有限公司"
        excel_b64 = _build_batch_excel(batch_name, units, farm_name, image_paths)

        # 清理临时文件
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)

        return {"success": True, "excel_base64": excel_b64}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"重新生成 Excel 失败: {exc}")


def _sanitize_excel_text(val: str) -> str:
    """防止 Excel 公式注入：以 = + - @ 开头的文本加单引号前缀。"""
    if val and val[0] in "=+-@":
        return "'" + val
    return val


def _parse_batch_folder(batch_name: str) -> dict:
    """从文件夹名解析栋舍和楼层，如 '育肥C区2楼1-4单元' → {'building': '育肥C区', 'floor': '2楼'}"""
    import re
    # 匹配完整栋舍名（含前缀）+ 楼层
    m = re.search(r'((?:育肥|保育|种猪)?.+?(?:区|舍))\s*(\d+楼)', batch_name)
    if m:
        return {"building": m.group(1).strip(), "floor": m.group(2).strip()}
    return {"building": batch_name, "floor": ""}


def _simplify_unit_name(unit_name: str) -> str:
    """去掉单元名中的栋舍楼层前缀，如 'C区2楼1单元' → '1单元'"""
    import re
    m = re.search(r'(\d+单元)$', unit_name)
    return m.group(1) if m else unit_name


def _extract_pen_number(pen_name: str) -> str:
    """从栏舍文件名提取数字，如 '10号栏舍.jpg' → '10'"""
    import re
    m = re.match(r'(\d+)', pen_name)
    return m.group(1) if m else pen_name


def _build_batch_excel(
    batch_name: str,
    units: list[dict],
    farm_name: str = "乐清市华统牧业有限公司",
    image_paths: list[str] | None = None,
) -> str:
    """基于模板生成 Excel 并返回 base64 字符串。"""
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.styles import Border, Side, Font, Alignment
    from copy import copy
    import re

    template_path = Path(__file__).parent / "excel_template.xlsx"
    wb = load_workbook(str(template_path))
    ws = wb.active

    folder_info = _parse_batch_folder(batch_name)

    # 0. 先取消所有合并，再删行，避免 openpyxl 内部状态冲突
    for mg in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mg))

    ws.cell(row=1, column=1, value="市本级活体生猪盘点一览表")
    ws.merge_cells('A1:J1')
    ws.delete_rows(2, 1)

    # 删除后行号偏移
    DATA_START = 3
    TEMPLATE_DATA_END = 10
    SUMMARY_ORIG = 11

    # 2. 清除数据区 {{占位符}} + 模板原始汇总行内容
    for r in range(DATA_START, SUMMARY_ORIG + 1):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and re.match(r'^\{\{.*\}\}$', cell.value):
                cell.value = None
    # 清除模板原始汇总行（第11行）的残留内容
    for c in range(1, 11):
        ws.cell(row=SUMMARY_ORIG, column=c).value = None
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, str) and re.match(r'^\{\{.*\}\}$', cell.value):
                cell.value = None

    # 2. 展开所有栏舍（简化名称）
    flat_pens = []
    total_pigs = 0
    unit_subtotals = []  # [(unit_name, pen_count, subtotal), ...]
    for unit in units:
        simple_name = _simplify_unit_name(unit["unit_name"])
        pen_count = len(unit["pens"])
        unit_subtotals.append((simple_name, pen_count, unit["subtotal"]))
        for pen in unit["pens"]:
            flat_pens.append({
                "unit_name": simple_name,
                "pen_number": _extract_pen_number(pen["pen_name"]),
                "pig_count": pen["pig_count"],
            })
            total_pigs += pen["pig_count"]

    data_count = len(flat_pens)

    # 3. 数据超过8行时插入额外行
    extra = data_count - 8
    if extra > 0:
        ws.insert_rows(SUMMARY_ORIG, amount=extra)
        for offset in range(extra):
            new_r = TEMPLATE_DATA_END + 1 + offset
            for c in range(1, 11):
                src = ws.cell(row=TEMPLATE_DATA_END, column=c)
                dst = ws.cell(row=new_r, column=c)
                dst.border = copy(src.border)
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)

    # 4. 数据区结束行
    data_end = DATA_START + data_count - 1

    # 5. 填充数据 + B/C/D 每行写入
    thin = Side(style="thin")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # 记录每个 unit 块的起止行
    cursor = DATA_START
    unit_row_ranges = []
    for unit_name, pen_count, subtotal in unit_subtotals:
        start_r = cursor
        end_r = cursor + pen_count - 1
        unit_row_ranges.append((start_r, end_r, unit_name, subtotal))
        cursor += pen_count

    for i, item in enumerate(flat_pens):
        r = DATA_START + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=_sanitize_excel_text(farm_name))
        ws.cell(row=r, column=3, value=_sanitize_excel_text(folder_info["building"]))
        ws.cell(row=r, column=4, value=_sanitize_excel_text(folder_info["floor"]))
        ws.cell(row=r, column=5, value=_sanitize_excel_text(item["unit_name"]))
        ws.cell(row=r, column=6, value=_sanitize_excel_text(item["pen_number"]))
        ws.cell(row=r, column=7, value=item["pig_count"])
        # 全框线
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = cell_border
            ws.cell(row=r, column=c).alignment = center

    # 6. E列（所在单元）+ I列（单元总数）按 unit 块合并
    for start_r, end_r, _, subtotal in unit_row_ranges:
        if end_r > start_r:
            ws.merge_cells(start_row=start_r, start_column=5, end_row=end_r, end_column=5)
            ws.merge_cells(start_row=start_r, start_column=9, end_row=end_r, end_column=9)
        ws.cell(row=start_r, column=9, value=subtotal)

    # 7. 汇总行：A-D 留空，E:H 合并显示"累计："，I 列显示总数
    summary_row = data_end + 1
    for c in range(1, 5):
        ws.cell(row=summary_row, column=c).value = None
    ws.merge_cells(start_row=summary_row, start_column=5, end_row=summary_row, end_column=8)
    ws.cell(row=summary_row, column=5, value="累计：")
    ws.cell(row=summary_row, column=9, value=total_pigs)
    for c in range(1, 11):
        ws.cell(row=summary_row, column=c).border = cell_border
        ws.cell(row=summary_row, column=c).alignment = center

    # 8. 插入原图到 H 列（Pillow 缩放到 350x210, quality 100）
    ws.column_dimensions['H'].width = 50
    ROW_HEIGHT_PT = 160
    IMG_TARGET_W, IMG_TARGET_H = 350, 210

    if image_paths:
        thumb_dir = Path(image_paths[0]).parent if image_paths else None
        for i, img_path in enumerate(image_paths):
            r = DATA_START + i
            if r > data_end:
                break
            ws.row_dimensions[r].height = ROW_HEIGHT_PT
            if img_path and Path(img_path).exists():
                try:
                    from PIL import Image as PilImage
                    pil_img = PilImage.open(img_path)
                    pil_img.thumbnail((IMG_TARGET_W, IMG_TARGET_H), PilImage.LANCZOS)
                    thumb_path = thumb_dir / f"thumb_{i}.jpg"
                    pil_img.save(str(thumb_path), quality=100)
                    xl_img = XlImage(str(thumb_path))
                    ws.add_image(xl_img, f'H{r}')
                except Exception:
                    pass

    buf = _BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host=DEFAULT_HOST, port=DEFAULT_PORT, reload=False)
